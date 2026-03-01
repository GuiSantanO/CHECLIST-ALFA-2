import customtkinter as ctk
from tkinter import messagebox, filedialog, ttk
import wmi
import psutil
import platform
import os
import sys
import subprocess
import datetime
import webbrowser
import threading
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# --- CONFIGURAÇÃO DE TEMA ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")  # Pode ser "blue" (padrão), "green", "dark-blue"

# Cores Personalizadas (Hex) - JANS-it Theme
COLOR_CARD = "#151515"     # Fundo dos Cartões/Frames (Quase preto)
COLOR_ACCENT = "#ff0000"   # Vermelho JANS-it (Ações Principais)
COLOR_TEXT = "#ffffff"     # Texto Principal (Branco)
COLOR_TEXT_DIM = "#b3b3b3" # Texto Secundário (Cinzento Claro)
COLOR_INFO = "#333333"     # Cinza Escuro (Botões Secundários)
PASSWORD_REGISTOS = "picanha2026" # Senha para acessar registos

# --- CONFIGURAÇÃO DE DIRETÓRIOS ---
# Define pasta de dados local (ao lado do script/executável) para portabilidade
# Se estiver congelado (PyInstaller), usa sys.executable, senão usa __file__
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR = os.path.join(BASE_DIR, "Dados")

if not os.path.exists(DATA_DIR):
    try:
        os.makedirs(DATA_DIR)
    except Exception as e:
        print(f"Erro ao criar diretório de dados: {e}")

EXCEL_FILE = os.path.join(DATA_DIR, "registos_checklist.xlsx")
EXCEL_FILE_TV = os.path.join(DATA_DIR, "registos_tvs.xlsx")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # Configuração da Janela
        self.title("Sistema de Checklist Pro")
        self.geometry("900x700")
        self.minsize(800, 600)
        
        # Cache de Info do Sistema
        self.sys_info = {}
        
        # Container Principal (para gestão de 'páginas')
        self.container = ctk.CTkFrame(self, fg_color="transparent")
        self.container.pack(fill="both", expand=True)
        
        # Inicializa Telas
        self.frames = {}
        self.current_frame = None
        
        # Carregar Hardware em Background
        threading.Thread(target=self.load_system_info_bg, daemon=True).start()
        
        # Mostrar Menu Inicial
        self.show_frame("MenuPrincipal")

    def load_system_info_bg(self):
        """Carrega info do sistema sem travar a UI"""
        try:
            # WMI precisa de inicialização COM em threads separadas
            import pythoncom
            pythoncom.CoInitialize()
        except:
            pass
            
        self.sys_info = get_system_info()
        
        # Agendar atualização na thread principal
        self.after(0, self.update_ui_after_load)

    def update_ui_after_load(self):
        """Atualiza a UI na thread principal"""
        if "ChecklistFrame" in self.frames:
            self.frames["ChecklistFrame"].update_hardware_info()

    def show_frame(self, page_name, direction="left"):
        """Alterna entre telas com uma transição suave de deslizar (slide)"""
        old_frame = self.current_frame
        
        # Cria a nova tela se ainda não existir ou recria para resetar
        if page_name == "MenuPrincipal":
            new_frame = MenuPrincipal(self.container, self)
        elif page_name == "ChecklistFrame":
            new_frame = ChecklistFrame(self.container, self)
        elif page_name == "ChecklistTVFrame":
            new_frame = ChecklistTVFrame(self.container, self)
        elif page_name == "RegistosFrame":
            new_frame = RegistosFrame(self.container, self)
        elif page_name == "RegistosTVFrame":
            new_frame = RegistosTVFrame(self.container, self)
        else:
            return

        self.current_frame = new_frame
        
        # Se for o primeiro frame a abrir (sem transição)
        if not old_frame:
            new_frame.place(relwidth=1, relheight=1, relx=0, rely=0)
            return

        # Configurar posições iniciais para animação
        start_x = 1.0 if direction == "left" else -1.0
        end_x = -1.0 if direction == "left" else 1.0
        
        new_frame.place(relwidth=1, relheight=1, relx=start_x, rely=0)
        
        # Iniciar loop de animação
        self.animate_transition(old_frame, new_frame, current_x=0.0, target_x=end_x, new_start_x=start_x)

    def animate_transition(self, old_frame, new_frame, current_x, target_x, new_start_x):
        """Atualiza iterativamente as coordenadas X com atenuação (Ease-Out) para um deslize suave"""
        # Calcular distância restante para atenuar a velocidade
        distance = abs(target_x - current_x)
        
        # O passo (step) diminui conforme a janela se aproxima do destino final, 
        # com um mínimo de 0.015 para garantir que a animação não para.
        step = max(distance * 0.15, 0.015)
        
        # Calcula direção do movimento
        moving_left = target_x < 0
        
        if moving_left:
            next_x = current_x - step
            new_x = new_start_x - step
        else:
            next_x = current_x + step
            new_x = new_start_x + step
            
        # Verifica se a animação terminou
        finished = (moving_left and next_x <= target_x) or (not moving_left and next_x >= target_x)
        
        if finished:
            new_frame.place(relx=0, rely=0)
            old_frame.place_forget()
            old_frame.destroy()
        else:
            old_frame.place(relx=next_x, rely=0)
            new_frame.place(relx=new_x, rely=0)
            new_start_x = new_x
            current_x = next_x
            # Agenda o próximo frame da animação (15ms ~ 60fps constantes)
            self.after(15, self.animate_transition, old_frame, new_frame, current_x, target_x, new_start_x)

# --- TELAS ---

class MenuPrincipal(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent, fg_color="transparent")
        self.controller = controller
        
        # Centralizar conteúdo
        self.center_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.center_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        # Título
        ctk.CTkLabel(self.center_frame, text="CHECKLIST RECONDICIONADOS", 
                    font=("Roboto Medium", 32), text_color=COLOR_TEXT).pack(pady=(0, 10))
        
        ctk.CTkLabel(self.center_frame, text="Selecione uma opção para continuar", 
                    font=("Roboto", 14), text_color=COLOR_TEXT_DIM).pack(pady=(0, 40))
        
        # Botões (Cartões Grandes)
        self.create_menu_button("🖥️  NOVA CHECKLIST PC", COLOR_ACCENT, 
                              lambda: controller.show_frame("ChecklistFrame"))
                              
        self.create_menu_button("📺  NOVA CHECKLIST TV", COLOR_ACCENT, 
                              lambda: controller.show_frame("ChecklistTVFrame"))
        
        self.create_menu_button("🗂️  ABRIR REGISTOS PC", "#3498db", 
                              self.check_password_registos)
                              
        self.create_menu_button("🗂️  ABRIR REGISTOS TV", "#1abc9c", 
                              self.check_password_registos_tv)
        
        self.create_menu_button("🔍  EXPORTAR DANOS", "#e67e22", 
                              exportar_danos_ui)
        
        self.create_menu_button("📊  EXPORTAR COMPRA (PDF)", "#8e44ad", 
                              exportar_compra_pdf_ui)

        # Rodapé

        ctk.CTkLabel(self, text="v3.0 - Design Minimalista", 
                    font=("Roboto", 10), text_color=COLOR_TEXT_DIM).pack(side="bottom", pady=20)

    def create_menu_button(self, text, color, command):
        # Separar o ícone do texto para estilizar cores independentemente
        partes = text.split("  ", 1)
        icon = partes[0]
        label_text = partes[1] if len(partes) > 1 else text

        btn = ctk.CTkButton(self.center_frame, text="", command=command,
                           fg_color="transparent", border_width=2, border_color=color,
                           hover_color=color, height=50, width=300,
                           corner_radius=25) # Botão arredondado
        btn.pack(pady=10)
        
        # Rótulo de Ícone (mantém a sua cor de emoji nativa do SO)
        lbl_icon = ctk.CTkLabel(btn, text=icon, font=("Segoe UI Emoji", 18), fg_color="transparent")
        lbl_icon.place(relx=0.15, rely=0.5, anchor="center")
        
        # Rótulo do Texto (sempre branco brilhante para contraste máximo contra #151515)
        lbl_text = ctk.CTkLabel(btn, text=label_text, text_color="#ffffff", font=("Roboto Medium", 14), fg_color="transparent")
        lbl_text.place(relx=0.25, rely=0.5, anchor="w")
        
        # Sincronizar Animação de Hover e Cliques pelas Camadas
        def on_enter(e): 
            btn.configure(fg_color=color)
            lbl_icon.configure(fg_color=color)
            lbl_text.configure(fg_color=color)
        def on_leave(e): 
            btn.configure(fg_color="transparent")
            lbl_icon.configure(fg_color="transparent")
            lbl_text.configure(fg_color="transparent")
        def on_click(e):
            command()
            
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        lbl_icon.bind("<Enter>", on_enter)
        lbl_icon.bind("<Leave>", on_leave)
        lbl_icon.bind("<Button-1>", on_click)
        
        lbl_text.bind("<Enter>", on_enter)
        lbl_text.bind("<Leave>", on_leave)
        lbl_text.bind("<Button-1>", on_click)

    def _ask_password(self, title):
        """Custom dialog para perguntar password com censura '*' """
        dialog = ctk.CTkToplevel(self)
        dialog.title(title)
        dialog.geometry("400x200")
        dialog.resizable(False, False)
        dialog.attributes("-topmost", True)
        
        # Tentar centrar o dialog
        dialog.update_idletasks()
        try:
            x = self.winfo_rootx() + (self.winfo_width() // 2) - 200
            y = self.winfo_rooty() + (self.winfo_height() // 2) - 100
            dialog.geometry(f"+{x}+{y}")
        except: pass
        
        dialog.grab_set()
        
        password = [None]
        
        lbl = ctk.CTkLabel(dialog, text="Digite a palavra-passe de administrador:", font=("Roboto", 14))
        lbl.pack(pady=(30, 10))
        
        entry = ctk.CTkEntry(dialog, show="*", width=250, font=("Roboto", 14))
        entry.pack(pady=(0, 20))
        entry.focus()
        
        def on_ok(_=None):
            password[0] = entry.get()
            dialog.destroy()
            
        def on_cancel(_=None):
            dialog.destroy()
            
        entry.bind("<Return>", on_ok)
        entry.bind("<Escape>", on_cancel)
        
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack()
        
        btn_ok = ctk.CTkButton(btn_frame, text="OK", width=100, command=on_ok)
        btn_ok.pack(side="left", padx=10)
        
        btn_cancel = ctk.CTkButton(btn_frame, text="Cancelar", width=100, fg_color="#e74c3c", hover_color="#c0392b", command=on_cancel)
        btn_cancel.pack(side="right", padx=10)
        
        dialog.wait_window()
        return password[0]

    def check_password_registos(self):
        """Solicita senha antes de abrir os registos"""
        password = self._ask_password("Acesso Restrito")
        
        if password == PASSWORD_REGISTOS:
            self.controller.show_frame("RegistosFrame")
        elif password is not None: # Se não cancelou
            messagebox.showerror("Erro", "Palavra-passe incorreta!")

    def check_password_registos_tv(self):
        """Solicita senha antes de abrir os registos de TV"""
        password = self._ask_password("Acesso Restrito (TVs)")
        
        if password == PASSWORD_REGISTOS:
            self.controller.show_frame("RegistosTVFrame")
        elif password is not None:
            messagebox.showerror("Erro", "Palavra-passe incorreta!")

class ChecklistFrame(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent, fg_color="transparent")
        self.controller = controller
        
        # Top Bar
        top_bar = ctk.CTkFrame(self, fg_color=COLOR_CARD, height=60, corner_radius=0)
        top_bar.pack(fill="x", side="top")
        
        ctk.CTkButton(top_bar, text="⬅ Voltar", command=lambda: controller.show_frame("MenuPrincipal", direction="right"),
                     fg_color="transparent", text_color=COLOR_TEXT, width=80).pack(side="left", padx=10)
        
        ctk.CTkLabel(top_bar, text="NOVO RELATÓRIO", font=("Roboto Medium", 18)).pack(side="left", padx=20)
        
        # Scroll Area Principal
        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=20, pady=20)
        
        # --- SEÇÃO 1: TÉCNICO ---
        self.add_section_header("1. Técnico Responsável")
        self.user_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.user_frame.pack(fill="x", pady=(0, 20))
        
        users = ["Guilherme", "Alex", "Araujo", "Convidado"]
        self.user_var = ctk.StringVar(value=users[0])
        self.combo_user = ctk.CTkComboBox(self.user_frame, values=users, variable=self.user_var,
                                        command=self.check_guest, width=300)
        self.combo_user.pack(padx=20, pady=20, anchor="w")
        
        self.entry_guest = ctk.CTkEntry(self.user_frame, placeholder_text="Nome do Convidado", width=300)
        
        # --- SEÇÃO 2: HARDWARE ---
        self.add_section_header("2. Hardware Detectado")
        self.hw_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.hw_frame.pack(fill="x", pady=(0, 20))
        
        self.lbl_model = ctk.CTkLabel(self.hw_frame, text="Detectando...", font=("Roboto Medium", 16))
        self.lbl_model.pack(padx=20, pady=(15, 5), anchor="w")
        
        self.lbl_specs = ctk.CTkLabel(self.hw_frame, text="Aguarde...", text_color=COLOR_TEXT_DIM, justify="left")
        self.lbl_specs.pack(padx=20, pady=(0, 15), anchor="w")
        
        self.update_hardware_info()
        
        # Detalhes de Memória adicionais removidos devido à deteção nativa mais inteligente
        
        # --- SEÇÃO 3: COMPRA ---
        self.add_section_header("3. Referência de Compra")
        self.compra_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.compra_frame.pack(fill="x", pady=(0, 20))
        
        self.entry_compra = ctk.CTkEntry(self.compra_frame, placeholder_text="Ex: 123456", width=300)
        self.entry_compra.pack(padx=20, pady=20, anchor="w")
        
        # --- SEÇÃO 4: TESTES ---
        self.add_section_header("4. Checklist de Testes")
        self.test_vars = {}
        tests = ["Teclado", "Ecrã", "Touch Screen", "Wifi", "LAN", "Portas USB", "Webcam", "Microfone", "Colunas", "Saídas Vídeo", "LTE"]
        
        self.tests_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.tests_frame.pack(fill="x", pady=(0, 20))
        
        # Grid para os testes
        # Grid para os testes
        for i, test in enumerate(tests):
            row = i // 2
            col = i % 2
            self.create_test_item(self.tests_frame, test, row, col)


        # --- SEÇÃO 5: NOTAS ---
        self.add_section_header("5. Observações e Danos")
        self.notes_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.notes_frame.pack(fill="x", pady=(0, 20))
        
        self.text_notes = ctk.CTkTextbox(self.notes_frame, height=100)
        self.text_notes.pack(fill="x", padx=20, pady=(20, 10))
        
        self.quick_notes_frame = ctk.CTkFrame(self.notes_frame, fg_color="transparent")
        self.quick_notes_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        quick_texts = [
            "pressure marks lcd", "broken", "lcd", "battery", "ko", "dead",
            "keyboard", "usb", "damaged", "palm rest", "domain locked"
        ]
        
        row, col = 0, 0
        for text in quick_texts:
            btn = ctk.CTkButton(self.quick_notes_frame, text=text, width=80,
                               fg_color=COLOR_INFO, hover_color="#2980b9",
                               command=lambda t=text: self.add_quick_note(t))
            btn.grid(row=row, column=col, padx=4, pady=4)
            col += 1
            if col > 4: # 5 botões por linha
                col = 0
                row += 1

        # --- BOTÃO AÇÃO ---
        self.btn_save = ctk.CTkButton(self.scroll, text="GERAR RELATÓRIO E GUARDAR", 
                                     fg_color=COLOR_ACCENT, hover_color="#cc0000",
                                     height=50, font=("Roboto Medium", 14),
                                     command=self.gerar_relatorio)
        self.btn_save.pack(fill="x", pady=20)

    def add_quick_note(self, text):
        current_text = self.text_notes.get("1.0", "end-1c").strip()
        if current_text:
            self.text_notes.insert("end", f", {text}")
        else:
            self.text_notes.insert("end", text)

    def open_incognito(self, url):
        """Abre o link no Edge InPrivate em Full Screen"""
        try:
            if sys.platform == 'win32':
                # Executa o Edge InPrivate e em Full Screen
                subprocess.run(f'start msedge --inprivate --start-fullscreen "{url}"', shell=True, check=True)
            else:
                webbrowser.open(url)
        except:
            # Fallback padrão
            webbrowser.open(url)

    def add_section_header(self, text):
        ctk.CTkLabel(self.scroll, text=text, font=("Roboto Medium", 14), 
                    text_color=COLOR_ACCENT).pack(anchor="w", pady=(10, 5))

    def check_guest(self, choice):
        if choice == "Convidado":
            self.entry_guest.pack(padx=20, pady=(0, 20), anchor="w")
        else:
            self.entry_guest.pack_forget()

    def update_hardware_info(self):
        info = self.controller.sys_info
        if info:
            self.lbl_model.configure(text=info.get('modelo', 'Desconhecido'))
            specs = f"S/N: {info.get('serial', 'N/A')}\nCPU: {info.get('cpu', 'N/A')}\nRAM: {info.get('ram', 'N/A')}\nDISCO: {info.get('disk', 'N/A')}\nGPU: {info.get('gpu', 'N/A')}\nECRÃ: {info.get('resolution', 'N/A')} ({info.get('refresh_rate', 'N/A')})"
            self.lbl_specs.configure(text=specs)

    def create_test_item(self, parent, test_name, row, col):
        # Frame individual para cada teste
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.grid(row=row, column=col, sticky="ew", padx=20, pady=10)
        
        # Label
        ctk.CTkLabel(frame, text=test_name, font=("Roboto", 12)).pack(anchor="w")
        
        # Switch Moderno
        var = ctk.BooleanVar(value=False)
        self.test_vars[test_name] = var
        
        switch = ctk.CTkSwitch(frame, text="Aprovado", variable=var, 
                              progress_color=COLOR_ACCENT, button_color="#ffffff")
        switch.pack(anchor="w")
        
        # Link opcional
        test_urls = {
            "Teclado": "https://en.key-test.ru/",
            "Webcam": "https://pt.webcamtests.com/",
            "Microfone": "https://pt.mictests.com/",
            "Colunas": "https://www.xbitlabs.com/pt/teste-de-som/",
            "Ecrã": "https://deadpixeltest.org/",
            "Touch Screen": "https://testmyscreen.com/"
        }
        if test_name in test_urls:
             link = ctk.CTkLabel(frame, text="Abrir Teste Online 🔗", text_color=COLOR_INFO, 
                                cursor="hand2", font=("Roboto", 10))
             link.pack(anchor="w")
             link.bind("<Button-1>", lambda e, u=test_urls[test_name]: self.open_incognito(u))

    def gerar_relatorio(self):
        # Coletar Dados
        usuario = self.user_var.get()
        if usuario == "Convidado":
            usuario = self.entry_guest.get() or "Convidado"
            
        compra_num = self.entry_compra.get().strip()
        
        # Validação do Nº Compra
        if not compra_num:
            messagebox.showwarning("Aviso", "A Referência de Compra é de preenchimento obrigatório.")
            return
            
        # Verifica se contém algo que não seja alfanumérico ou barra '/'
        if not compra_num.replace('/', '').replace('\\', '').isalnum():
            messagebox.showwarning("Aviso", "A Referência de Compra só pode conter letras, números e barras (/).\nCaracteres especiais não são permitidos.")
            return
            
        danos = self.text_notes.get("1.0", "end-1c")
        
        if getattr(self.controller, 'sys_info', None) is None:
            messagebox.showerror("Aviso", "Aguarde a deteção do hardware.")
            return
            
        testes = {name: var.get() for name, var in self.test_vars.items()}
        
        # Mapeamento de nomes para o Excel/HTML (para compatibilidade com código antigo)
        testes_mapped = {
            "Teclado": testes.get("Teclado"),
            "Ecrã": testes.get("Ecrã"),
            "Touch Screen": testes.get("Touch Screen"),
            "Wifi": testes.get("Wifi"),
            "LAN": testes.get("LAN"),
            "Webcam": testes.get("Webcam"),
            "Microfone": testes.get("Microfone"),
            "Colunas": testes.get("Colunas"),
            "USB": testes.get("Portas USB"), 
            "Portas de Vídeo": testes.get("Saídas Vídeo"),
            "LTE": testes.get("LTE")
        }
        
        # Chaman lógica de geração (reutilizando função externa refatorada ou movida)
        gerar_relatorio_logic(self.controller.sys_info, usuario, compra_num, testes_mapped, danos)

class ChecklistTVFrame(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent, fg_color="transparent")
        self.controller = controller
        
        # Top Bar
        top_bar = ctk.CTkFrame(self, fg_color=COLOR_CARD, height=60, corner_radius=0)
        top_bar.pack(fill="x", side="top")
        
        ctk.CTkButton(top_bar, text="⬅ Voltar", command=lambda: controller.show_frame("MenuPrincipal", direction="right"),
                     fg_color="transparent", text_color=COLOR_TEXT, width=80).pack(side="left", padx=10)
        
        ctk.CTkLabel(top_bar, text="NOVO RELATÓRIO TV", font=("Roboto Medium", 18)).pack(side="left", padx=20)
        
        # Scroll Area Principal
        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=20, pady=20)
        
        # --- SEÇÃO 1: TÉCNICO ---
        self.add_section_header("1. Técnico Responsável")
        self.user_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.user_frame.pack(fill="x", pady=(0, 20))
        
        users = ["Guilherme", "Alex", "Araujo", "Convidado"]
        self.user_var = ctk.StringVar(value=users[0])
        self.combo_user = ctk.CTkComboBox(self.user_frame, values=users, variable=self.user_var,
                                        command=self.check_guest, width=300)
        self.combo_user.pack(padx=20, pady=20, anchor="w")
        
        self.entry_guest = ctk.CTkEntry(self.user_frame, placeholder_text="Nome do Convidado", width=300)
        
        # --- SEÇÃO 2: HARDWARE TV ---
        self.add_section_header("2. TV Detectada")
        self.hw_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.hw_frame.pack(fill="x", pady=(0, 20))
        
        self.entry_model = ctk.CTkEntry(self.hw_frame, font=("Roboto Medium", 16), width=400)
        self.entry_model.pack(padx=20, pady=(15, 5), anchor="w")
        
        self.lbl_specs = ctk.CTkLabel(self.hw_frame, text="Aguarde...", text_color=COLOR_TEXT_DIM, justify="left")
        self.lbl_specs.pack(padx=20, pady=(0, 15), anchor="w")
        
        self.btn_detect = ctk.CTkButton(self.hw_frame, text="Detectar Ecrã Novamente", command=self.update_hardware_info, width=200)
        self.btn_detect.pack(padx=20, pady=(0, 15), anchor="w")
        
        self.tv_info = None
        self.update_hardware_info()
        
        # --- SEÇÃO 3: COMPRA ---
        self.add_section_header("3. Referência de Compra")
        self.compra_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.compra_frame.pack(fill="x", pady=(0, 20))
        
        self.entry_compra = ctk.CTkEntry(self.compra_frame, placeholder_text="Ex: 123456", width=300)
        self.entry_compra.pack(padx=20, pady=20, anchor="w")
        
        # --- SEÇÃO 4: TESTES ---
        self.add_section_header("4. Checklist de Testes")
        self.test_vars = {}
        tests = ["Ecrã / Imagem", "Touch Screen", "Colunas", "Cabos / Energia", "Botões", "Comando (Remote)", "Webcam"]
        
        self.tests_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.tests_frame.pack(fill="x", pady=(0, 20))
        
        for i, test in enumerate(tests):
            row = i // 2
            col = i % 2
            self.create_test_item(self.tests_frame, test, row, col)

        # --- SEÇÃO 5: PORTAS ---
        self.add_section_header("5. Portas de Vídeo (Quantidade)")
        self.port_vars = {}
        ports = ["DisplayPort", "HDMI", "DVI", "VGA", "RS232", "USB", "USB A", "USB C"]
        
        self.ports_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.ports_frame.pack(fill="x", pady=(0, 20))
        
        for i, port in enumerate(ports):
            row = i // 3
            col = i % 3
            self.create_port_item(self.ports_frame, port, row, col)

        # --- SEÇÃO 6: NOTAS ---
        self.add_section_header("6. Observações e Danos")
        self.notes_frame = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD)
        self.notes_frame.pack(fill="x", pady=(0, 20))
        
        self.text_notes = ctk.CTkTextbox(self.notes_frame, height=100)
        self.text_notes.pack(fill="x", padx=20, pady=(20, 10))
        
        self.quick_notes_frame = ctk.CTkFrame(self.notes_frame, fg_color="transparent")
        self.quick_notes_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        quick_texts = [
            "riscos no ecrã", "sem comando", "botão preso", "dead pixels", "sem cabo",
            "carcaça partida", "mancha no lcd"
        ]
        
        row, col = 0, 0
        for text in quick_texts:
            btn = ctk.CTkButton(self.quick_notes_frame, text=text, width=80,
                               fg_color=COLOR_INFO, hover_color="#2980b9",
                               command=lambda t=text: self.add_quick_note(t))
            btn.grid(row=row, column=col, padx=4, pady=4)
            col += 1
            if col > 4: 
                col = 0
                row += 1

        # --- BOTÃO AÇÃO ---
        self.btn_save = ctk.CTkButton(self.scroll, text="GERAR RELATÓRIO E GUARDAR", 
                                     fg_color=COLOR_ACCENT, hover_color="#cc0000",
                                     height=50, font=("Roboto Medium", 14),
                                     command=self.gerar_relatorio)
        self.btn_save.pack(fill="x", pady=20)

    def check_guest(self, choice):
        if choice == "Convidado":
            self.entry_guest.pack(padx=20, pady=(0, 20), anchor="w")
        else:
            self.entry_guest.pack_forget()

    def update_hardware_info(self):
        self.tv_info = get_tv_info()
        if self.tv_info:
            specs = f"S/N: {self.tv_info.get('serial', 'N/A')}\nResolução: {self.tv_info.get('resolution', 'N/A')}\nRefresh Rate: {self.tv_info.get('refresh_rate', 'N/A')}"
            self.lbl_specs.configure(text=specs)
            
            auto_modelo = f"{self.tv_info.get('marca', 'Marca Desconhecida')} - {self.tv_info.get('modelo', 'Modelo Desconhecido')}"
            self.entry_model.delete(0, "end")
            self.entry_model.insert(0, auto_modelo)

    def create_test_item(self, parent, test_name, row, col):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.grid(row=row, column=col, sticky="ew", padx=20, pady=10)
        ctk.CTkLabel(frame, text=test_name, font=("Roboto", 12)).pack(anchor="w")
        var = ctk.BooleanVar(value=False)
        self.test_vars[test_name] = var
        switch = ctk.CTkSwitch(frame, text="Aprovado", variable=var, 
                              progress_color=COLOR_ACCENT, button_color="#ffffff")
        switch.pack(anchor="w")
        
    def create_port_item(self, parent, port_name, row, col):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.grid(row=row, column=col, sticky="ew", padx=20, pady=10)
        
        # Checkbox + quantity entry
        var_check = ctk.BooleanVar(value=False)
        var_qty = ctk.StringVar(value="0")
        
        def toggle_entry():
            if var_check.get():
                entry.configure(state="normal")
                if var_qty.get() == "0":
                    var_qty.set("1")
            else:
                var_qty.set("0")
                entry.configure(state="disabled")
                
        chk = ctk.CTkCheckBox(frame, text=port_name, variable=var_check, command=toggle_entry)
        chk.pack(anchor="w", side="top", pady=(0,5))
        
        entry = ctk.CTkEntry(frame, textvariable=var_qty, width=50, state="disabled")
        entry.pack(anchor="w", side="left")
        
        self.port_vars[port_name] = var_qty

    def add_quick_note(self, text):
        current_text = self.text_notes.get("1.0", "end-1c").strip()
        if current_text:
            self.text_notes.insert("end", f", {text}")
        else:
            self.text_notes.insert("end", text)

    def gerar_relatorio(self):
        usuario = self.user_var.get()
        if usuario == "Convidado":
            usuario = self.entry_guest.get() or "Convidado"
            
        compra_num = self.entry_compra.get().strip()
        
        if not compra_num:
            messagebox.showwarning("Aviso", "A Referência de Compra é de preenchimento obrigatório.")
            return
            
        if not compra_num.replace('/', '').replace('\\', '').isalnum():
            messagebox.showwarning("Aviso", "A Referência de Compra só pode conter letras, números e barras (/).\nCaracteres especiais não são permitidos.")
            return
            
        danos = self.text_notes.get("1.0", "end-1c")
        
        # Override tv_info marca and model with the user's manual entry if edited
        marca_modelo = self.entry_model.get().strip()
        self.tv_info['marca'] = ""
        self.tv_info['modelo'] = marca_modelo
        
        testes = {name: var.get() for name, var in self.test_vars.items()}
        
        portas = {}
        for name, var in self.port_vars.items():
            try:
                qtd = int(var.get())
                if qtd > 0:
                    portas[name] = qtd
            except ValueError:
                pass
                
        gerar_relatorio_tv_logic(self.tv_info, usuario, compra_num, portas, testes, danos)

    def add_section_header(self, text):
        ctk.CTkLabel(self.scroll, text=text, font=("Roboto Medium", 14), 
                    text_color=COLOR_ACCENT).pack(anchor="w", pady=(10, 5))

class RegistosFrame(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent, fg_color="transparent")
        self.controller = controller
        
        # Top Bar
        top_bar = ctk.CTkFrame(self, fg_color=COLOR_CARD, height=60, corner_radius=0)
        top_bar.pack(fill="x", side="top")
        
        ctk.CTkButton(top_bar, text="⬅ Voltar", command=lambda: controller.show_frame("MenuPrincipal", direction="right"),
                     fg_color="transparent", text_color=COLOR_TEXT, width=80).pack(side="left", padx=10)
        
        ctk.CTkLabel(top_bar, text="REGISTOS DE CHECKLIST", font=("Roboto Medium", 18)).pack(side="left", padx=20)
        
        # Main Layout
        self.main_area = ctk.CTkFrame(self, fg_color="transparent")
        self.main_area.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Right Panel (Editor) - Packed first so it takes priority on the right side
        self.editor_frame = ctk.CTkFrame(self.main_area, fg_color=COLOR_CARD, width=300)
        self.editor_frame.pack(side="right", fill="y", padx=(15, 0))
        self.editor_frame.pack_propagate(False) # Força a manter os 300px de largura
        
        # Left Panel (Table) - Fills the rest of the available space
        self.tree_frame = ctk.CTkFrame(self.main_area, fg_color=COLOR_CARD)
        self.tree_frame.pack(side="left", fill="both", expand=True)
        
        # Search and Filter Bar
        self.filter_frame = ctk.CTkFrame(self.tree_frame, fg_color="transparent")
        self.filter_frame.pack(fill="x", padx=10, pady=(10, 0))
        
        self.search_var = ctk.StringVar()
        self.search_entry = ctk.CTkEntry(self.filter_frame, placeholder_text="Procurar (S/N, Modelo, Nº Compra...)", 
                                        textvariable=self.search_var, width=300)
        self.search_entry.pack(side="left", padx=(0, 10))
        self.search_var.trace_add("write", lambda *args: self.apply_filters())
        
        self.sort_var = ctk.StringVar(value="Data")
        self.sort_combo = ctk.CTkComboBox(self.filter_frame, values=["Data", "Ordem Alfabética (Modelo)"], 
                                         variable=self.sort_var, command=lambda e: self.apply_filters(), width=200)
        self.sort_combo.pack(side="right")
        
        ctk.CTkLabel(self.filter_frame, text="Ordenar por:").pack(side="right", padx=(0, 10))
        
        # Style
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", 
                        background=COLOR_CARD,
                        foreground=COLOR_TEXT,
                        rowheight=25,
                        fieldbackground=COLOR_CARD,
                        bordercolor=COLOR_CARD,
                        borderwidth=0)
        style.map('Treeview', background=[('selected', COLOR_INFO)])
        style.configure("Treeview.Heading",
                        background="#333333",
                        foreground=COLOR_TEXT,
                        relief="flat")
        style.map("Treeview.Heading", background=[('active', "#444444")])

        # Scrollbars
        self.tree_scroll_y = ctk.CTkScrollbar(self.tree_frame, orientation="vertical")
        self.tree_scroll_y.pack(side="right", fill="y")
        
        self.tree_scroll_x = ctk.CTkScrollbar(self.tree_frame, orientation="horizontal")
        self.tree_scroll_x.pack(side="bottom", fill="x")
        
        self.tree = ttk.Treeview(self.tree_frame, yscrollcommand=self.tree_scroll_y.set, xscrollcommand=self.tree_scroll_x.set, selectmode="extended")
        self.tree.pack(fill="both", expand=True)
        
        self.tree_scroll_y.configure(command=self.tree.yview)
        self.tree_scroll_x.configure(command=self.tree.xview)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        

        
        ctk.CTkLabel(self.editor_frame, text="✏️ Editor", font=("Roboto Medium", 16)).pack(pady=10)
        
        self.editor_scroll = ctk.CTkScrollableFrame(self.editor_frame, fg_color="transparent")
        self.editor_scroll.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.btn_save = ctk.CTkButton(self.editor_frame, text="GUARDAR ALTERAÇÕES", 
                                     fg_color=COLOR_ACCENT, hover_color="#cc0000",
                                     command=self.save_edits)
        self.btn_save.pack(pady=15, padx=10, fill="x")
        
        self.df = None
        self.filtered_df = None
        self.current_idx = None
        self.editor_widgets = {}
        
        # Load Data on init
        self.after(100, self.load_data)

    def load_data(self):
        if not os.path.exists(EXCEL_FILE):
             self.tree.insert("", "end", values=("Nenhum registo de momento.",))
             return
             
        try:
             self.df = pd.read_excel(EXCEL_FILE)
             if 'Tipo RAM' in self.df.columns: self.df = self.df.drop(columns=['Tipo RAM'])
             if 'Config RAM' in self.df.columns: self.df = self.df.drop(columns=['Config RAM'])
             self.apply_filters()
             self.create_editor_fields()
        except Exception as e:
             messagebox.showerror("Erro", f"Falha ao ler Excel: {e}")

    def apply_filters(self):
        if self.df is None or self.df.empty: return
        
        query = self.search_var.get().strip().lower()
        sort_by = self.sort_var.get()
        
        # Copiar dataframe original
        self.filtered_df = self.df.copy()
        
        # Converter colunas necessárias para string para evitar erros de busca
        for col in ['Modelo', 'Serial', 'Nº Compra']:
             if col in self.filtered_df.columns:
                 self.filtered_df[col] = self.filtered_df[col].astype(str)
                 
        # Aplicar Pesquisa
        if query:
            # Procurar em Modelo, Serial, CPU, GPU, Nº Compra
            mask = self.filtered_df.apply(lambda row: any(query in str(val).lower() for val in row), axis=1)
            self.filtered_df = self.filtered_df[mask]
            
        # Aplicar Ordenação
        if sort_by == "Data":
            # Tentar ordenar por data real se possível, senão pelo index inverso (mais recente primeiro se adicionado no fim)
            try:
                self.filtered_df['DataRaw'] = pd.to_datetime(self.filtered_df['Data'], format="%d/%m/%Y %H:%M", errors='coerce')
                self.filtered_df = self.filtered_df.sort_values(by='DataRaw', ascending=False)
                self.filtered_df = self.filtered_df.drop(columns=['DataRaw'])
            except:
                self.filtered_df = self.filtered_df.sort_index(ascending=False)
        elif sort_by == "Ordem Alfabética (Modelo)":
            if 'Modelo' in self.filtered_df.columns:
                self.filtered_df = self.filtered_df.sort_values(by='Modelo', ascending=True)

        self.populate_tree()
        
    def populate_tree(self):
         # Reset tree
         for item in self.tree.get_children():
             self.tree.delete(item)
             
         if self.filtered_df is None or self.filtered_df.empty:
             return
             
         # Setup columns
         self.tree["columns"] = list(self.df.columns)
         self.tree["show"] = "headings"
         
         for col in self.df.columns:
             self.tree.heading(col, text=col)
             width = 150 if col in ["Modelo", "Serial", "Notas", "CPU", "Disco", "GPU"] else 80
             self.tree.column(col, width=width, minwidth=50, stretch=False, anchor="center")
             
         # Insert rows
         for idx, row in self.filtered_df.iterrows():
             values = []
             for col_name, val in row.items():
                 if pd.isna(val) or str(val) == "nan":
                     values.append("")
                 else:
                     val_str = str(val)
                     if val_str.endswith('.0') and col_name in ["Nº Compra"]:
                         val_str = val_str[:-2]
                     values.append(val_str)
             self.tree.insert("", "end", iid=str(idx), values=values)

    def create_editor_fields(self):
        if self.df is None or self.df.empty: return
        
        for w in self.editor_scroll.winfo_children():
            w.destroy()
        self.editor_widgets.clear()
        
        read_only_cols = ["Data"]
        
        for col in self.df.columns:
            lbl = ctk.CTkLabel(self.editor_scroll, text=col, font=("Roboto", 12))
            lbl.pack(anchor="w", pady=(5, 0))
            
            if col == "Notas":
                entry = ctk.CTkTextbox(self.editor_scroll, height=80)
                entry.pack(fill="x", pady=2)
            elif col in ["Teclado", "Ecrã", "Touch Screen", "Wifi", "LAN", "Webcam", "Microfone", "Colunas", "USB", "Portas de Vídeo", "LTE"]:
                entry = ctk.CTkComboBox(self.editor_scroll, values=["✓", "✗", "N/A"])
                entry.pack(fill="x", pady=2)
            else:
                entry = ctk.CTkEntry(self.editor_scroll)
                entry.pack(fill="x", pady=2)
            
            if col in read_only_cols and hasattr(entry, 'configure'):
                try: entry.configure(state="disabled")
                except: pass
                
            self.editor_widgets[col] = entry

    def on_tree_select(self, event):
        selected = self.tree.selection()
        if not selected:
             # Retirado o disable do btn e o clear caso perca o foco por clicar num Entry
             return
             
        self.current_idx = int(selected[0])
        
        item_values = self.tree.item(selected[0], "values")
        
        for i, col in enumerate(self.df.columns):
             widget = self.editor_widgets.get(col)
             if widget:
                 was_disabled = False
                 try:
                     was_disabled = (widget.cget("state") == "disabled")
                 except Exception:
                     pass
                 
                 if was_disabled: 
                     try: widget.configure(state="normal")
                     except: pass
                 
                 val = item_values[i]
                 
                 if isinstance(widget, ctk.CTkTextbox):
                     widget.delete("1.0", "end")
                     widget.insert("1.0", val if val != "nan" else "")
                 elif isinstance(widget, ctk.CTkComboBox):
                     widget.set(val if val != "nan" else "")
                 else:
                     widget.delete(0, "end")
                     widget.insert(0, val if val != "nan" else "")
                     
                 if was_disabled: 
                     try: widget.configure(state="disabled")
                     except: pass

    def save_edits(self):
        if self.current_idx is None:
             messagebox.showwarning("Aviso", "Por favor, selecione um registo na tabela à esquerda antes de guardar alterações.")
             return
        if self.df is None: return
        
        try:
             # Converter current_idx para int explicitamente por segurança, embora já devesse vir como int do on_tree_select
             idx = int(self.current_idx)
             
             for col, widget in self.editor_widgets.items():
                 is_disabled = False
                 try:
                     is_disabled = (widget.cget("state") == "disabled")
                 except Exception:
                     pass
                     
                 if is_disabled:
                     continue
                     
                 if isinstance(widget, ctk.CTkTextbox):
                     val = widget.get("1.0", "end-1c")
                 else:
                     val = widget.get()
                     
                 # Evitar erro de dtype int64/float64 ao guardar strings (ex. N.º de Série muito grande)
                 if str(self.df[col].dtype) != 'object':
                     self.df[col] = self.df[col].astype(object)
                     
                 self.df.at[idx, col] = val
                 
             self.df.to_excel(EXCEL_FILE, index=False, sheet_name="Registos")
             formatar_excel(EXCEL_FILE)
             
             fazer_backup(EXCEL_FILE, "PC")
             
             self.load_data()
             messagebox.showinfo("Sucesso", "Registo atualizado e guardado!")
        except Exception as e:
              messagebox.showerror("Erro", f"Erro ao guardar: {e}")

class RegistosTVFrame(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent, fg_color="transparent")
        self.controller = controller
        
        top_bar = ctk.CTkFrame(self, fg_color=COLOR_CARD, height=60, corner_radius=0)
        top_bar.pack(fill="x", side="top")
        
        ctk.CTkButton(top_bar, text="⬅ Voltar", command=lambda: controller.show_frame("MenuPrincipal", direction="right"),
                     fg_color="transparent", text_color=COLOR_TEXT, width=80).pack(side="left", padx=10)
        
        ctk.CTkLabel(top_bar, text="REGISTOS DE TVs", font=("Roboto Medium", 18)).pack(side="left", padx=20)
        
        self.main_area = ctk.CTkFrame(self, fg_color="transparent")
        self.main_area.pack(fill="both", expand=True, padx=20, pady=20)
        
        self.editor_frame = ctk.CTkFrame(self.main_area, fg_color=COLOR_CARD, width=300)
        self.editor_frame.pack(side="right", fill="y", padx=(15, 0))
        self.editor_frame.pack_propagate(False)
        
        self.tree_frame = ctk.CTkFrame(self.main_area, fg_color=COLOR_CARD)
        self.tree_frame.pack(side="left", fill="both", expand=True)
        
        self.filter_frame = ctk.CTkFrame(self.tree_frame, fg_color="transparent")
        self.filter_frame.pack(fill="x", padx=10, pady=(10, 0))
        
        self.search_var = ctk.StringVar()
        self.search_entry = ctk.CTkEntry(self.filter_frame, placeholder_text="Procurar (S/N, Marca, Modelo, Nº Compra...)", 
                                        textvariable=self.search_var, width=300)
        self.search_entry.pack(side="left", padx=(0, 10))
        self.search_var.trace_add("write", lambda *args: self.apply_filters())
        
        self.sort_var = ctk.StringVar(value="Data")
        self.sort_combo = ctk.CTkComboBox(self.filter_frame, values=["Data", "Ordem Alfabética (Modelo)"], 
                                         variable=self.sort_var, command=lambda e: self.apply_filters(), width=200)
        self.sort_combo.pack(side="right")
        
        ctk.CTkLabel(self.filter_frame, text="Ordenar por:").pack(side="right", padx=(0, 10))
        
        style = ttk.Style()
        style.theme_use("default")
        style.configure("TV.Treeview", 
                        background=COLOR_CARD,
                        foreground=COLOR_TEXT,
                        rowheight=25,
                        fieldbackground=COLOR_CARD,
                        bordercolor=COLOR_CARD,
                        borderwidth=0)
        style.map('TV.Treeview', background=[('selected', COLOR_INFO)])
        style.configure("TV.Treeview.Heading",
                        background="#333333",
                        foreground=COLOR_TEXT,
                        relief="flat")
        style.map("TV.Treeview.Heading", background=[('active', "#444444")])

        self.tree_scroll_y = ctk.CTkScrollbar(self.tree_frame, orientation="vertical")
        self.tree_scroll_y.pack(side="right", fill="y")
        
        self.tree_scroll_x = ctk.CTkScrollbar(self.tree_frame, orientation="horizontal")
        self.tree_scroll_x.pack(side="bottom", fill="x")
        
        self.tree = ttk.Treeview(self.tree_frame, style="TV.Treeview", yscrollcommand=self.tree_scroll_y.set, xscrollcommand=self.tree_scroll_x.set, selectmode="extended")
        self.tree.pack(fill="both", expand=True)
        
        self.tree_scroll_y.configure(command=self.tree.yview)
        self.tree_scroll_x.configure(command=self.tree.xview)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        
        ctk.CTkLabel(self.editor_frame, text="✏️ Editor", font=("Roboto Medium", 16)).pack(pady=10)
        
        self.editor_scroll = ctk.CTkScrollableFrame(self.editor_frame, fg_color="transparent")
        self.editor_scroll.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.btn_save = ctk.CTkButton(self.editor_frame, text="GUARDAR ALTERAÇÕES", 
                                     fg_color=COLOR_ACCENT, hover_color="#cc0000",
                                     command=self.save_edits)
        self.btn_save.pack(pady=15, padx=10, fill="x")
        
        self.df = None
        self.filtered_df = None
        self.current_idx = None
        self.editor_widgets = {}
        
        self.after(100, self.load_data)

    def load_data(self):
        if not os.path.exists(EXCEL_FILE_TV):
             self.tree.insert("", "end", values=("Nenhum registo de TVs de momento.",))
             return
        try:
             self.df = pd.read_excel(EXCEL_FILE_TV)
             
             if 'Marca' in self.df.columns and 'Modelo' in self.df.columns:
                 self.df['Marca'] = self.df['Marca'].fillna("")
                 self.df['Modelo'] = self.df['Modelo'].fillna("")
                 marca = self.df['Marca'].astype(str).str.replace('nan', '', case=False)
                 modelo = self.df['Modelo'].astype(str).str.replace('nan', '', case=False)
                 self.df.insert(3, 'Marca/Modelo', (marca + " " + modelo).str.strip())
                 self.df.loc[self.df['Marca/Modelo'] == "", 'Marca/Modelo'] = "Desconhecido"
                 self.df = self.df.drop(columns=['Marca', 'Modelo'])
                 self.df.to_excel(EXCEL_FILE_TV, index=False, sheet_name="RegistosTV")
                 formatar_excel_tv(EXCEL_FILE_TV)
                 
             self.apply_filters()
             self.create_editor_fields()
        except Exception as e:
             messagebox.showerror("Erro", f"Falha ao ler Excel: {e}")

    def apply_filters(self):
        if self.df is None or self.df.empty: return
        query = self.search_var.get().strip().lower()
        sort_by = self.sort_var.get()
        self.filtered_df = self.df.copy()
        
        for col in ['Marca/Modelo', 'Serial', 'Nº Compra']:
             if col in self.filtered_df.columns:
                 self.filtered_df[col] = self.filtered_df[col].astype(str)
                 
        if query:
            mask = self.filtered_df.apply(lambda row: any(query in str(val).lower() for val in row), axis=1)
            self.filtered_df = self.filtered_df[mask]
            
        if sort_by == "Data":
            try:
                self.filtered_df['DataRaw'] = pd.to_datetime(self.filtered_df['Data'], format="%d/%m/%Y %H:%M", errors='coerce')
                self.filtered_df = self.filtered_df.sort_values(by='DataRaw', ascending=False)
                self.filtered_df = self.filtered_df.drop(columns=['DataRaw'])
            except:
                self.filtered_df = self.filtered_df.sort_index(ascending=False)
        elif sort_by == "Ordem Alfabética (Modelo)":
            if 'Marca/Modelo' in self.filtered_df.columns:
                self.filtered_df = self.filtered_df.sort_values(by='Marca/Modelo', ascending=True)

        self.populate_tree()
        
    def populate_tree(self):
         for item in self.tree.get_children():
             self.tree.delete(item)
         if self.filtered_df is None or self.filtered_df.empty:
             return
         self.tree["columns"] = list(self.df.columns)
         self.tree["show"] = "headings"
         for col in self.df.columns:
             self.tree.heading(col, text=col)
             width = 150 if col in ["Marca/Modelo", "Serial", "Notas", "Resolução"] else 80
             self.tree.column(col, width=width, minwidth=50, stretch=False, anchor="center")
         for idx, row in self.filtered_df.iterrows():
             values = []
             for col_name, val in row.items():
                 if pd.isna(val) or str(val) == "nan":
                     values.append("")
                 else:
                     val_str = str(val)
                     if val_str.endswith('.0') and col_name in ["DisplayPort", "HDMI", "DVI", "VGA", "RS232", "USB", "USB A", "USB C", "Nº Compra"]:
                         val_str = val_str[:-2]
                     values.append(val_str)
             self.tree.insert("", "end", iid=str(idx), values=values)

    def create_editor_fields(self):
        if self.df is None or self.df.empty: return
        for w in self.editor_scroll.winfo_children(): w.destroy()
        self.editor_widgets.clear()
        
        read_only_cols = ["Data"]
        for col in self.df.columns:
            lbl = ctk.CTkLabel(self.editor_scroll, text=col, font=("Roboto", 12))
            lbl.pack(anchor="w", pady=(5, 0))
            
            if col == "Notas":
                entry = ctk.CTkTextbox(self.editor_scroll, height=80)
                entry.pack(fill="x", pady=2)
            elif col in ["Ecrã / Imagem", "Touch Screen", "Colunas", "Cabos / Energia", "Botões", "Comando (Remote)", "Webcam"]:
                entry = ctk.CTkComboBox(self.editor_scroll, values=["✓", "✗", "N/A"])
                entry.pack(fill="x", pady=2)
            else:
                entry = ctk.CTkEntry(self.editor_scroll)
                entry.pack(fill="x", pady=2)
            
            if col in read_only_cols and hasattr(entry, 'configure'):
                try: entry.configure(state="disabled")
                except: pass
            self.editor_widgets[col] = entry

    def on_tree_select(self, event):
        selected = self.tree.selection()
        if not selected: return
             
        self.current_idx = int(selected[0])
        item_values = self.tree.item(selected[0], "values")
        
        for i, col in enumerate(self.df.columns):
             widget = self.editor_widgets.get(col)
             if widget:
                 was_disabled = False
                 try: was_disabled = (widget.cget("state") == "disabled")
                 except: pass
                 if was_disabled: 
                     try: widget.configure(state="normal")
                     except: pass
                 
                 val = item_values[i]
                 if isinstance(widget, ctk.CTkTextbox):
                     widget.delete("1.0", "end")
                     widget.insert("1.0", val if val != "nan" else "")
                 elif isinstance(widget, ctk.CTkComboBox):
                     widget.set(val if val != "nan" else "")
                 else:
                     widget.delete(0, "end")
                     
                     if val.endswith('.0') and col in ["DisplayPort", "HDMI", "DVI", "VGA", "RS232", "USB", "USB A", "USB C"]:
                         val = val[:-2]
                         
                     widget.insert(0, val if val != "nan" else "")
                     
                 if was_disabled: 
                     try: widget.configure(state="disabled")
                     except: pass

    def save_edits(self):
        if self.current_idx is None:
             messagebox.showwarning("Aviso", "Por favor, selecione um registo na tabela antes.")
             return
        if self.df is None: return
        try:
             idx = int(self.current_idx)
             for col, widget in self.editor_widgets.items():
                 is_disabled = False
                 try: is_disabled = (widget.cget("state") == "disabled")
                 except: pass
                 if is_disabled: continue
                     
                 if isinstance(widget, ctk.CTkTextbox):
                     val = widget.get("1.0", "end-1c")
                 else:
                     val = widget.get()
                     
                 if str(self.df[col].dtype) != 'object':
                     self.df[col] = self.df[col].astype(object)
                     
                 self.df.at[idx, col] = val
                 
             self.df.to_excel(EXCEL_FILE_TV, index=False, sheet_name="RegistosTV")
             formatar_excel_tv(EXCEL_FILE_TV)
             
             fazer_backup(EXCEL_FILE_TV, "TV")
             
             self.load_data()
             messagebox.showinfo("Sucesso", "Registo atualizado e guardado!")
        except Exception as e:
             messagebox.showerror("Erro", f"Erro ao guardar TV: {e}")



# --- LÓGICA DO SISTEMA (ADAPTADA) ---

def get_system_info():
    info = {}
    try:
        c = wmi.WMI()
        # 1. Modelo
        try:
            sys_data = c.Win32_ComputerSystem()[0]
            info['modelo'] = f"{sys_data.Manufacturer} {sys_data.Model}".strip()
        except: info['modelo'] = "Modelo Desconhecido"
            
        # 2. Serial (Melhorado - Prioriza SystemProduct > BIOS > BaseBoard)
        try:
            # Tenta pegar o "Service Tag" ou Identificador do Sistema (Mais confiável)
            product = c.Win32_ComputerSystemProduct()[0]
            serial = product.IdentifyingNumber.strip()
            
            # Filtra valores genéricos comuns
            if serial.lower() in ["default string", "system serial number", "to be filled by o.e.m.", "0"]:
                raise ValueError("Serial genérico")
            
            info['serial'] = serial
        except:
            try:
                # Tenta BIOS (Muito comum em laptops)
                bios = c.Win32_Bios()[0]
                serial = bios.SerialNumber.strip()
                
                if serial.lower() in ["default string", "system serial number", "to be filled by o.e.m.", "0"]:
                    raise ValueError("Serial genérico")
                
                info['serial'] = serial
            except:
                try:
                     # Último recurso: BaseBoard (Pode ser o serial da placa mãe interna)
                    board = c.Win32_BaseBoard()[0]
                    info['serial'] = board.SerialNumber.strip()
                except: 
                    info['serial'] = "N/A"
        
        # 3. CPU
        try:
            proc = c.Win32_Processor()[0]
            info['cpu'] = proc.Name.strip()
        except: info['cpu'] = platform.processor()
        
        # 4. RAM
        try:
            total_ram = 0
            stick_specs = []
            for mem_stick in c.Win32_PhysicalMemory():
                cap_gb = round(int(mem_stick.Capacity) / (1024**3))
                total_ram += int(mem_stick.Capacity)
                
                speed_str = f" {mem_stick.Speed} MT/s" if mem_stick.Speed else ""
                
                # Check FormFactor for soldered RAM types
                # 12 commonly SODIMM (sometimes soldered in modern ultra-thins but not guaranteed)
                # 21, 22, 23, 24 represent specifically soldered form factors: BGA, FB-DIMM, etc.
                is_soldered = False
                if mem_stick.FormFactor in [21, 22, 23, 24]:
                    is_soldered = True
                else:
                    # OEMs frequently misreport soldered RAM as FormFactor 12. 
                    # Scan strings for explicit OEM labels
                    locator = str(getattr(mem_stick, 'DeviceLocator', '')).lower()
                    part = str(getattr(mem_stick, 'PartNumber', '')).lower()
                    if any(kw in locator for kw in ['onboard', 'on board', 'soldered', 'bga']):
                        is_soldered = True
                    elif any(kw in part for kw in ['onboard', 'on board', 'soldered', 'bga']):
                        is_soldered = True
                        
                soldered_str = " [Soldada]" if is_soldered else ""
                stick_specs.append(f"{cap_gb}GB{speed_str}{soldered_str}".strip())
                    
            ram_gb = round(total_ram / (1024**3))
            
            ram_str = f"{ram_gb} GB"
            if stick_specs:
                ram_str += f" ({' + '.join(stick_specs)})"
                
            info['ram'] = ram_str
                
        except: 
            try:
                ram_bytes = psutil.virtual_memory().total
                info['ram'] = f"{round(ram_bytes / (1024**3))} GB"
            except:
                info['ram'] = "N/A"
        
        # 5. Disco
        disks = []
        try:
            for disk in c.Win32_DiskDrive():
                size_gb = round(int(disk.Size) / (1024**3))
                disks.append(f"{disk.Model} ({size_gb} GB)")
            info['disk'] = " + ".join(disks)
        except: info['disk'] = "N/A"
        
        # 6. GPU & Resolution Info
        gpus = []
        try:
            res_width = ""
            res_height = ""
            refresh = ""
            
            for gpu in c.Win32_VideoController():
                gpus.append(gpu.Name)
                if gpu.CurrentHorizontalResolution and gpu.CurrentVerticalResolution:
                     res_width = gpu.CurrentHorizontalResolution
                     res_height = gpu.CurrentVerticalResolution
                     refresh = gpu.CurrentRefreshRate or "60" # Default se não detetar
                     
            info['gpu'] = " | ".join(gpus) if gpus else "N/A"
            info['resolution'] = f"{res_width}x{res_height}" if res_width else "N/A"
            info['refresh_rate'] = f"{refresh} Hz" if refresh else "N/A"
        except Exception as e:
            info['gpu'] = "N/A"
            info['resolution'] = "N/A"
            info['refresh_rate'] = "N/A"
        
    except Exception as e:
        info['error'] = str(e)
        return None
    return info

def gerar_relatorio_logic(sys_info, usuario, compra_num, testes, danos):
    """Gera o HTML e salva no Excel"""
    
    # HTML Content (Design Melhorado)
    foto_user = f"{usuario.lower()}.jpg"
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Relatório {sys_info.get('modelo', 'PC')}</title>
        <style>
            body {{ font-family: 'Segoe UI', sans-serif; margin: 0; background: #f4f7f6; color: #333; }}
            .container {{ max_width: 800px; margin: 40px auto; background: white; padding: 40px; border-radius: 12px; box-shadow: 0 10px 30px rgba(0,0,0,0.1); }}
            .header {{ display: flex; align-items: center; border-bottom: 2px solid #eee; padding-bottom: 20px; margin-bottom: 30px; }}
            .header img {{ width: 80px; height: 80px; border-radius: 50%; object-fit: cover; margin-right: 20px; border: 3px solid #eee; }}
            .header h1 {{ margin: 0; font-size: 24px; color: #2c3e50; }}
            .header p {{ margin: 5px 0 0; color: #7f8c8d; }}
            h2 {{ color: #2980b9; font-size: 18px; border-left: 4px solid #2980b9; padding-left: 10px; margin-top: 30px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ padding: 12px 15px; border-bottom: 1px solid #eee; text-align: left; }}
            th {{ background-color: #f8f9fa; color: #2c3e50; font-weight: 600; }}
            .pass {{ color: #27ae60; font-weight: bold; }}
            .fail {{ color: #e74c3c; font-weight: bold; }}
            .notes {{ background: #fff8e1; padding: 20px; border-radius: 8px; border-left: 4px solid #ffc107; margin-top: 15px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <img src="{foto_user}" onerror="this.src='https://ui-avatars.com/api/?name={usuario}&background=random'">
                <div>
                    <h1>Relatório Técnico de Certificação</h1>
                    <p>Técnico: <strong>{usuario}</strong> &bull; {datetime.datetime.now().strftime("%d/%m/%Y")}</p>
                </div>
            </div>

            <h2>📦 Identificação do Equipamento</h2>
            <table>
                <tr><th width="30%">Ref. Compra</th><td>{compra_num}</td></tr>
                <tr><th>Modelo</th><td><strong>{sys_info.get('modelo', 'N/A')}</strong></td></tr>
                <tr><th>Serial Number</th><td>{sys_info.get('serial', 'N/A')}</td></tr>
            </table>

            <h2>⚙️ Especificações Técnicas</h2>
            <table>
                <tr><th width="30%">Processador</th><td>{sys_info.get('cpu', 'N/A')}</td></tr>
                <tr><th>Memória RAM</th><td>{sys_info.get('ram', 'N/A')}</td></tr>
                <tr><th>Armazenamento</th><td>{sys_info.get('disk', 'N/A')}</td></tr>
                <tr><th>Gráfica</th><td>{sys_info.get('gpu', 'N/A')}</td></tr>
                <tr><th>Ecrã (Resolução)</th><td>{sys_info.get('resolution', 'N/A')}</td></tr>
                <tr><th>Ecrã (Taxa Atualização)</th><td>{sys_info.get('refresh_rate', 'N/A')}</td></tr>
            </table>

            <h2>✅ Resultados dos Testes</h2>
            <table>
                {''.join([f"<tr><td>{k}</td><td class='{'pass' if v else 'fail'}'>{'APROVADO' if v else 'REPROVADO'}</td></tr>" for k,v in testes.items()])}
            </table>

            <h2>📝 Observações</h2>
            <div class="notes">
                {danos.replace('\\n', '<br>') if danos and danos.strip() else "Nenhuma anomalia visual detetada. Equipamento em condições normais."}
            </div>
        </div>
    </body>
    </html>
    """
    
    # Salvar HTML
    safe_serial = "".join([c for c in sys_info.get('serial', 'SN') if c.isalnum()]).strip()
    default_filename = f"{safe_serial}.html"
    
    file_path = filedialog.asksaveasfilename(
        defaultextension=".html",
        filetypes=[("Ficheiros HTML", "*.html")],
        initialfile=default_filename,
        title="Guardar Relatório Como..."
    )
    
    if not file_path: return

    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(html_content)
            
        webbrowser.open('file://' + os.path.realpath(file_path))
        
        # Salvar Excel
        if guardar_em_excel(usuario, compra_num, sys_info, testes, danos):
            messagebox.showinfo("Sucesso", "Processo concluído com sucesso!")
        else:
            messagebox.showwarning("Atenção", "Relatório HTML gerado, mas erro ao salvar no Excel.")
            
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao gravar: {e}")

def fazer_backup(filepath, tipo):
    if not os.path.exists(filepath):
        return
    backup_dir = os.path.join(DATA_DIR, "backups", tipo)
    if not os.path.exists(backup_dir):
        try:
            os.makedirs(backup_dir)
        except Exception as e:
            print(f"Erro ao criar diretório de backup: {e}")
            return
            
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.basename(filepath)
    name, ext = os.path.splitext(filename)
    backup_filename = f"{name}_{timestamp}{ext}"
    backup_path = os.path.join(backup_dir, backup_filename)
    
    try:
        shutil.copy2(filepath, backup_path)
    except Exception as e:
        print(f"Erro ao fazer backup: {e}")

def guardar_em_excel(usuario, compra_num, sys_info, testes, danos):
    try:
        registo = {
            "Data": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
            "Técnico": usuario,
            "Nº Compra": compra_num,
            "Modelo": sys_info.get('modelo', 'N/A'),
            "Serial": sys_info.get('serial', 'N/A'),
            "CPU": sys_info.get('cpu', 'N/A'),
            "RAM": sys_info.get('ram', 'N/A'),
            "Disco": sys_info.get('disk', 'N/A'),
            "GPU": sys_info.get('gpu', 'N/A'),
            "Resolução": sys_info.get('resolution', 'N/A'),
            "Refresh": sys_info.get('refresh_rate', 'N/A'),
            "Teclado": "✓" if testes.get("Teclado") else "✗",
            "Ecrã": "✓" if testes.get("Ecrã") else "✗",
            "Touch Screen": "✓" if testes.get("Touch Screen") else "✗",
            "Wifi": "✓" if testes.get("Wifi") else "✗",
            "LAN": "✓" if testes.get("LAN") else "✗",
            "Webcam": "✓" if testes.get("Webcam") else "✗",
            "Microfone": "✓" if testes.get("Microfone") else "✗",
            "Colunas": "✓" if testes.get("Colunas") else "✗",
            "USB": "✓" if testes.get("USB") else "✗",
            "Portas de Vídeo": "✓" if testes.get("Portas de Vídeo") else "✗",
            "LTE": "✓" if testes.get("LTE") else "✗",
            "Notas": danos.strip() if danos.strip() else "Sem observações"
        }
        
        if os.path.exists(EXCEL_FILE):
             df_exist = pd.read_excel(EXCEL_FILE)
             df = pd.concat([df_exist, pd.DataFrame([registo])], ignore_index=True)
        else:
             df = pd.DataFrame([registo])
             
        # Forçar ordem das colunas
        cols_order = [
            "Data", "Técnico", "Nº Compra", "Modelo", "Serial", "CPU", "RAM",
            "Disco", "GPU", "Resolução", "Refresh",
            "Teclado", "Ecrã", "Touch Screen", "Wifi", "LAN", "Webcam", "Microfone", "Colunas", "USB", 
            "Portas de Vídeo", "LTE", "Notas"
        ]
        
        # Garantir que todas as colunas existem (se o excel antigo não tiver alguma)
        for col in cols_order:
            if col not in df.columns:
                df[col] = "N/A"
                
        # Reordenar
        df = df[cols_order]
             
        df.to_excel(EXCEL_FILE, index=False, sheet_name="Registos")
        formatar_excel(EXCEL_FILE)
        
        fazer_backup(EXCEL_FILE, "PC")
        return True
    except Exception as e:
        print(e)
        return False

def formatar_excel(filepath):
    """Aplica formatação moderna ao ficheiro Excel"""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Definir estilos
        header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid") # Vermelho JANS-it
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='B4C7E7'),
            right=Side(style='thin', color='B4C7E7'),
            top=Side(style='thin', color='B4C7E7'),
            bottom=Side(style='thin', color='B4C7E7')
        )
        
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Altura do cabeçalho
        ws.row_dimensions[1].height = 25
        
        # Formatar linhas de dados
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 2):
            # Cor alternada para as linhas
            fill = alt_fill if (row_idx % 2 == 0) else white_fill
            
            for col_idx, cell in enumerate(row, 1):
                cell.fill = fill
                cell.border = thin_border
                
                # Alinhar colunas de testes (marcas ✓ e ✗)
                if col_idx > 11:  # Colunas dos testes (ajustado para novas colunas e refresh/res)
                    cell.alignment = center_alignment
                    cell.font = Font(size=12, bold=True)
                    
                    # Colorir as marcas
                    if cell.value == "✓":
                        cell.font = Font(size=12, bold=True, color="00B050")
                    elif cell.value == "✗":
                        cell.font = Font(size=12, bold=True, color="C00000")
                else:
                    cell.alignment = left_alignment
                    cell.font = Font(size=10)
                
                ws.row_dimensions[row_idx].height = 20
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 15,  # Data
            'B': 10,  # Técnico
            'C': 10,  # Nº Compra
            'D': 25,  # Modelo
            'E': 15,  # Serial
            'F': 30,  # CPU
            'G': 35,  # RAM (Widened to fit explicit stick data)
            'H': 15,  # Disco
            'I': 20,  # GPU
            'J': 15,  # Resolução
            'K': 10,  # Refresh Rate
            'L': 8,   # Teclado
            'M': 8,   # Ecrã
            'N': 8,   # Touch Screen
            'O': 8,   # Wifi
            'P': 8,   # LAN
            'Q': 8,   # Webcam
            'R': 8,   # Microfone
            'S': 8,   # Colunas
            'T': 8,   # USB
            'U': 8,   # Portas de Vídeo
            'V': 8,   # LTE
            'W': 30   # Notas
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Congelar a primeira linha (cabeçalho)
        ws.freeze_panes = "A2"
        
        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Erro ao formatar Excel: {e}")
        return False

def formatar_excel_danos(filepath):
    """Formata o Excel de Danos para ficar estético e ajustado"""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Estilos
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid") # Dark Grey
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='B4C7E7'),
            right=Side(style='thin', color='B4C7E7'),
            top=Side(style='thin', color='B4C7E7'),
            bottom=Side(style='thin', color='B4C7E7')
        )
        
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Formatar Cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            
        ws.row_dimensions[1].height = 25
        
        # Formatar Dados e Ajustar Colunas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = left_alignment
                
                # Se for a coluna Serial (B), forçar formato de texto
                if cell.column_letter == 'B':
                    cell.number_format = '@'
                
                # Se for a coluna de Notas (C), permitir quebra de linha
                if cell.column_letter == 'C': 
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Ajustar larguras
        column_widths = {'A': 30, 'B': 25, 'C': 60}
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            if column == 'C':
                ws.column_dimensions[column].width = 60
                continue
                
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            
            adjusted_width = (max_length + 2)
            if adjusted_width < 10: adjusted_width = 10
            if adjusted_width > 50: adjusted_width = 50
            
            ws.column_dimensions[column].width = adjusted_width

        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Erro ao formatar Excel de danos: {e}")
        return False

def exportar_danos_ui():
    # Verificar se existe pelo menos um ficheiro
    if not os.path.exists(EXCEL_FILE) and not os.path.exists(EXCEL_FILE_TV):
        messagebox.showinfo("Aviso", "Ainda não existem registos para exportar.")
        return

    # Interface simples para input
    dialog = ctk.CTkInputDialog(text="Digite o Nº de Compra para exportar:", title="Exportar Danos")
    compra_num = dialog.get_input()
    
    if not compra_num: return
    
    try:
        df_final = pd.DataFrame()
        
        # 1. Procurar em PCs
        if os.path.exists(EXCEL_FILE):
            df_pcs = pd.read_excel(EXCEL_FILE)
            df_pcs['Nº Compra'] = df_pcs['Nº Compra'].astype(str)
            filtro_pcs = (df_pcs['Nº Compra'] == str(compra_num)) & (df_pcs['Notas'].notna()) & (df_pcs['Notas'] != "Sem observações")
            df_export_pcs = df_pcs[filtro_pcs][['Modelo', 'Serial', 'Notas']]
            if not df_export_pcs.empty:
                df_export_pcs.insert(0, 'Equipamento', 'PC')
                df_final = pd.concat([df_final, df_export_pcs], ignore_index=True)
                
        # 2. Procurar em TVs/Monitores
        if os.path.exists(EXCEL_FILE_TV):
            df_tvs = pd.read_excel(EXCEL_FILE_TV)
            df_tvs['Nº Compra'] = df_tvs['Nº Compra'].astype(str)
            filtro_tvs = (df_tvs['Nº Compra'] == str(compra_num)) & (df_tvs['Notas'].notna()) & (df_tvs['Notas'] != "Sem observações")
            
            # Create "Modelo" column depending on whether they used separated Marca/Modelo or combined "Marca/Modelo"
            if 'Marca/Modelo' in df_tvs.columns:
                df_temp = df_tvs[filtro_tvs][['Marca/Modelo', 'Serial', 'Notas']].rename(columns={'Marca/Modelo': 'Modelo'})
            elif 'Marca' in df_tvs.columns and 'Modelo' in df_tvs.columns:
                df_tvs['TempModelo'] = df_tvs['Marca'].astype(str) + " " + df_tvs['Modelo'].astype(str)
                df_temp = df_tvs[filtro_tvs][['TempModelo', 'Serial', 'Notas']].rename(columns={'TempModelo': 'Modelo'})
            else:
                df_temp = pd.DataFrame() # Fallback if columns are broken
                
            if not df_temp.empty:
                df_temp.insert(0, 'Equipamento', 'TV/Monitor')
                df_final = pd.concat([df_final, df_temp], ignore_index=True)
        
        if df_final.empty:
            messagebox.showinfo("Vazio", "Nenhum registo com danos encontrado para esta compra.")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=f"Danos_{compra_num}.xlsx")
        if save_path:
            df_final.to_excel(save_path, index=False)
            formatar_excel_danos(save_path)
            os.startfile(save_path)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")

def formatar_excel_compra_pdf(filepath):
    """Formata o Excel temporário para ficar apresentável no PDF com duas tabelas"""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Ocultar linhas de grelha
        ws.sheet_view.showGridLines = False
        
        header_fill_pc = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid") # Red
        header_fill_tv = PatternFill(start_color="333333", end_color="333333", fill_type="solid") # Dark Grey
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, color="333333", size=14)
        
        thin_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
                             top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Procurar onde começam os cabeçalhos das tabelas
        pc_start = None
        tv_start = None
        
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val == "COMPUTADORES":
                ws.cell(row=row_idx, column=1).font = title_font
                pc_start = row_idx + 1 # O cabeçalho real é na linha seguinte
            elif val == "MONITORES / TVs":
                ws.cell(row=row_idx, column=1).font = title_font
                tv_start = row_idx + 1
        
        # Formatar PC Table
        if pc_start:
            # Cabeçalho PC
            ws.merge_cells(start_row=pc_start, start_column=9, end_row=pc_start, end_column=12)
            for col in range(1, 13):
                cell = ws.cell(row=pc_start, column=col)
                if cell.value or col >= 9:
                    cell.fill = header_fill_pc
                    cell.font = header_font
                    cell.alignment = center_align
                cell.border = thin_border
            ws.row_dimensions[pc_start].height = 25
            
            # Dados PC
            r = pc_start + 1
            while r <= ws.max_row and ws.cell(row=r, column=1).value and ws.cell(row=r, column=1).value != "MONITORES / TVs":
                ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
                for col in range(1, 13):
                    c = ws.cell(row=r, column=col)
                    c.border = thin_border
                    c.alignment = center_align if col in [7, 8] else left_align # Refresh e Res center
                    if col == 2: c.number_format = '@' # Serial
                    if col == 9: c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True) # Notas
                r += 1
                
        # Formatar TV Table
        if tv_start:
            tv_cols = 12
            
            # Cabeçalho TV
            for col in range(1, tv_cols + 1):
                cell = ws.cell(row=tv_start, column=col)
                cell.fill = header_fill_tv
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            ws.row_dimensions[tv_start].height = 25
            
            # Dados TV
            r = tv_start + 1
            while r <= ws.max_row and (ws.cell(row=r, column=1).value or ws.cell(row=r, column=2).value):
                for col in range(1, tv_cols + 1):
                    c = ws.cell(row=r, column=col)
                    c.border = thin_border
                    c.alignment = center_align if (2 < col < 12) else left_align # Center all ports
                    if col == 2: c.number_format = '@' # Serial
                    if col == tv_cols: c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True) # Notas
                r += 1

        # Ajustar larguras globais (baseadas no PC que é mais largo, TV vai herdar)
        ws.column_dimensions['A'].width = 30 # Modelo/Marca
        ws.column_dimensions['B'].width = 18 # Serial
        ws.column_dimensions['C'].width = 22 # CPU / Resolução
        ws.column_dimensions['D'].width = 12 # RAM / Refresh
        ws.column_dimensions['E'].width = 18 # Disco / DP
        ws.column_dimensions['F'].width = 24 # GPU / HDMI
        ws.column_dimensions['G'].width = 12 # Res / DVI
        ws.column_dimensions['H'].width = 10 # Ref / VGA
        ws.column_dimensions['I'].width = 8  # PC Obs(m) / RS232
        ws.column_dimensions['J'].width = 8  # PC Obs(m) / USB
        ws.column_dimensions['K'].width = 8  # PC Obs(m) / USB C
        ws.column_dimensions['L'].width = 32 # PC Obs(m) / Notas TV
        
        # Assegurar que printa correto em folha inteira
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = False
        ws.page_setup.fitToWidth = 1
        
        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Erro formatando PDF Excel: {e}")
        return False

def exportar_compra_pdf_ui():
    if not os.path.exists(EXCEL_FILE) and not os.path.exists(EXCEL_FILE_TV):
        messagebox.showinfo("Aviso", "Ainda não existem registos para exportar.")
        return

    dialog = ctk.CTkInputDialog(text="Digite o Nº de Compra para exportar:", title="Exportar Compra")
    compra_num = dialog.get_input()
    
    if not compra_num: return
    
    export_ans = messagebox.askquestion("Formato de Exportação", "Deseja exportar o relatório final em formato PDF?\n\n(Sim = PDF | Não = Excel)", icon='question')
    is_pdf = (export_ans == 'yes')
    
    try:
        import win32com.client
        
        df_pcs_export = pd.DataFrame()
        if os.path.exists(EXCEL_FILE):
            df_pcs = pd.read_excel(EXCEL_FILE)
            df_pcs['Nº Compra'] = df_pcs['Nº Compra'].astype(str)
            filtro_pcs = df_pcs['Nº Compra'] == str(compra_num)
            colunas_pcs = ['Modelo', 'Serial', 'CPU', 'RAM', 'Disco', 'GPU', 'Resolução', 'Refresh', 'Notas']
            df_pcs_export = df_pcs[filtro_pcs].reindex(columns=colunas_pcs)
            df_pcs_export.rename(columns={'Notas': 'Observações'}, inplace=True)
            
        df_tvs_export = pd.DataFrame()
        if os.path.exists(EXCEL_FILE_TV):
            df_tvs = pd.read_excel(EXCEL_FILE_TV)
            df_tvs['Nº Compra'] = df_tvs['Nº Compra'].astype(str)
            filtro_tvs = df_tvs['Nº Compra'] == str(compra_num)
            if not df_tvs.empty:
                if 'Marca' in df_tvs.columns and 'Modelo' in df_tvs.columns:
                    df_tvs['Marca'] = df_tvs['Marca'].fillna("")
                    df_tvs['Modelo'] = df_tvs['Modelo'].fillna("")
                    marca = df_tvs['Marca'].astype(str).str.replace('nan', '', case=False)
                    modelo = df_tvs['Modelo'].astype(str).str.replace('nan', '', case=False)
                    df_tvs['Marca/Modelo'] = (marca + " " + modelo).str.strip()
                    df_tvs.loc[df_tvs['Marca/Modelo'] == "", 'Marca/Modelo'] = "Desconhecido"
                
                colunas_tvs = ['Marca/Modelo', 'Serial', 'Resolução', 'Refresh', 'DisplayPort', 'HDMI', 'DVI', 'VGA', 'RS232', 'USB', 'USB C', 'Notas']
                df_tvs_export = df_tvs[filtro_tvs].reindex(columns=colunas_tvs)
                df_tvs_export.rename(columns={'Notas': 'Observações'}, inplace=True)
        
        if df_pcs_export.empty and df_tvs_export.empty:
            messagebox.showinfo("Vazio", "Nenhum registo encontrado para esta compra.")
            return

        if is_pdf:
            default_name = f"Compra_{compra_num}.pdf"
            save_path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=default_name, filetypes=[("Documentos PDF", "*.pdf")])
        else:
            default_name = f"Compra_{compra_num}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name, filetypes=[("Folha Excel", "*.xlsx")])
            
        if not save_path: return
        
        # If PDF, we need a temp excel to format. If Excel, we just build it exactly where they asked to save.
        target_excel_path = os.path.join(DATA_DIR, f"temp_compra_{compra_num}.xlsx") if is_pdf else save_path
        
        # Write both dataframes to the same sheet with spacing
        with pd.ExcelWriter(target_excel_path, engine='openpyxl') as writer:
            start_row = 0
            
            if not df_pcs_export.empty:
                pd.DataFrame([["COMPUTADORES"]]).to_excel(writer, index=False, header=False, startrow=start_row)
                start_row += 1
                df_pcs_export.to_excel(writer, index=False, startrow=start_row)
                start_row += len(df_pcs_export) + 3 # Space between tables
                
            if not df_tvs_export.empty:
                pd.DataFrame([["MONITORES / TVs"]]).to_excel(writer, index=False, header=False, startrow=start_row)
                start_row += 1
                
                # Replace port '.0' decimals for cleaner PDF/Excel
                for col in ['DisplayPort', 'HDMI', 'DVI', 'VGA', 'RS232', 'USB', 'USB C']:
                    if col in df_tvs_export.columns:
                        df_tvs_export[col] = df_tvs_export[col].astype(str).str.replace(r'\.0$', '', regex=True).replace('nan', '')
                        
                df_tvs_export.to_excel(writer, index=False, startrow=start_row)

        formatar_excel_compra_pdf(target_excel_path)
        
        if is_pdf:
            try:
                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                
                wb = excel.Workbooks.Open(os.path.abspath(target_excel_path))
                
                ws = wb.ActiveSheet
                ws.PageSetup.Orientation = 2 # xlLandscape
                ws.PageSetup.Zoom = False
                ws.PageSetup.FitToPagesWide = 1
                ws.PageSetup.FitToPagesTall = False
                
                wb.ExportAsFixedFormat(0, os.path.abspath(save_path))
                wb.Close(False)
                excel.Quit()
                
                # Apagar temp file
                if os.path.exists(target_excel_path):
                    os.remove(target_excel_path)
                    
            except Exception as e:
                messagebox.showerror("Erro PDF", f"Falha na conversão para PDF: {e}")
                return
                
        os.startfile(save_path)
        messagebox.showinfo("Sucesso", "Ficheiro exportado com sucesso!")
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")



def get_tv_info():
    info = {
        'modelo': 'Modelo Desconhecido',
        'marca': 'Marca Desconhecida',
        'serial': 'N/A',
        'resolution': 'N/A',
        'refresh_rate': 'N/A'
    }
    try:
        c = wmi.WMI()
        import win32api
        res_width = ""
        res_height = ""
        refresh = ""
        
        try:
            monitors = win32api.EnumDisplayMonitors()
            target_device = None
            if len(monitors) > 1:
                for m in monitors:
                    info = win32api.GetMonitorInfo(m[0])
                    if info.get('Flags') != 1:
                        target_device = info.get('Device')
                        break
            if not target_device and monitors:
                target_device = win32api.GetMonitorInfo(monitors[0][0]).get('Device')
            if target_device:
                settings = win32api.EnumDisplaySettings(target_device, -1)
                res_width = settings.PelsWidth
                res_height = settings.PelsHeight
                refresh = settings.DisplayFrequency
        except Exception as e:
            print("win32api err:", e)
            for gpu in c.Win32_VideoController():
                if gpu.CurrentHorizontalResolution and gpu.CurrentVerticalResolution:
                     res_width = gpu.CurrentHorizontalResolution
                     res_height = gpu.CurrentVerticalResolution
                     refresh = gpu.CurrentRefreshRate or "60"
                     break
                     
        if res_width and res_height:
            info['resolution'] = f"{res_width}x{res_height}"
            
        if refresh:
            info['refresh_rate'] = f"{refresh} Hz"
            
        # Marca, Modelo e Serial
        try:
            cw = wmi.WMI(namespace="root\\wmi")
            # Vai iterar os monitores
            for m in cw.WmiMonitorID():
                if m.Active:
                    user_friendly = "".join([chr(char) for char in m.UserFriendlyName if char != 0]) if m.UserFriendlyName else ""
                    manufacturer = "".join([chr(char) for char in m.ManufacturerName if char != 0]) if m.ManufacturerName else ""
                    product = "".join([chr(char) for char in m.ProductCodeID if char != 0]) if m.ProductCodeID else ""
                    serial = "".join([chr(char) for char in m.SerialNumberID if char != 0]) if m.SerialNumberID else ""
                    
                    if user_friendly:
                        info['modelo'] = user_friendly
                    elif product:
                        info['modelo'] = product
                        
                    if manufacturer:
                        info['marca'] = manufacturer
                        
                    if serial:
                        info['serial'] = serial
                    break # Pega primeiro ativo
        except Exception as e:
            print("WmiMonitorID err:", e)

    except Exception as e:
        info['error'] = str(e)
        return None
    return info

def gerar_relatorio_tv_logic(sys_info, usuario, compra_num, portas, testes, danos):
    """Gera o HTML para a TV e salva no Excel"""
    
    foto_user = f"{usuario.lower()}.jpg"
    
    portas_html = "".join([f"<tr><th>{k}</th><td>{v}</td></tr>" for k,v in portas.items() if v > 0])
    if not portas_html:
         portas_html = "<tr><th>Nenhuma porta selecionada</th><td>-</td></tr>"
         
    testes_html = "".join([f"<tr><td>{k}</td><td class='{'pass' if v else 'fail'}'>{'APROVADO' if v else 'REPROVADO'}</td></tr>" for k,v in testes.items()])
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Relatório Monitor/TV {sys_info.get('marca', '')} {sys_info.get('modelo', 'TV')}</title>
        <style>
            body {{ font-family: 'Segoe UI', sans-serif; margin: 0; background: #f4f7f6; color: #333; }}
            .container {{ max-width: 800px; margin: 40px auto; background: white; padding: 40px; border-radius: 12px; box-shadow: 0 10px 30px rgba(0,0,0,0.1); }}
            .header {{ display: flex; align-items: center; border-bottom: 2px solid #eee; padding-bottom: 20px; margin-bottom: 30px; }}
            .header img {{ width: 80px; height: 80px; border-radius: 50%; object-fit: cover; margin-right: 20px; border: 3px solid #eee; }}
            .header h1 {{ margin: 0; font-size: 24px; color: #2c3e50; }}
            .header p {{ margin: 5px 0 0; color: #7f8c8d; }}
            h2 {{ color: #e67e22; font-size: 18px; border-left: 4px solid #e67e22; padding-left: 10px; margin-top: 30px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ padding: 12px 15px; border-bottom: 1px solid #eee; text-align: left; }}
            th {{ background-color: #f8f9fa; color: #2c3e50; font-weight: 600; }}
            .pass {{ color: #27ae60; font-weight: bold; }}
            .fail {{ color: #e74c3c; font-weight: bold; }}
            .notes {{ background: #fff8e1; padding: 20px; border-radius: 8px; border-left: 4px solid #ffc107; margin-top: 15px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <img src="{foto_user}" onerror="this.src='https://ui-avatars.com/api/?name={usuario}&background=random'">
                <div>
                    <h1>Relatório Técnico Monitor/TV</h1>
                    <p>Técnico: <strong>{usuario}</strong> &bull; {datetime.datetime.now().strftime("%d/%m/%Y")}</p>
                </div>
            </div>

            <h2>📦 Identificação do Equipamento</h2>
            <table>
                <tr><th width="30%">Ref. Compra</th><td>{compra_num}</td></tr>
                <tr><th>Marca / Modelo</th><td><strong>{sys_info.get('marca', 'N/A')} {sys_info.get('modelo', 'N/A')}</strong></td></tr>
                <tr><th>Serial Number</th><td>{sys_info.get('serial', 'N/A')}</td></tr>
            </table>

            <h2>🖥️ Especificações do Ecrã</h2>
            <table>
                <tr><th width="30%">Resolução</th><td>{sys_info.get('resolution', 'N/A')}</td></tr>
                <tr><th>Taxa Atualização</th><td>{sys_info.get('refresh_rate', 'N/A')}</td></tr>
            </table>

            <h2>🔌 Portas de Vídeo / Conexões (Qtd)</h2>
            <table>
                {portas_html}
            </table>

            <h2>✅ Resultados dos Testes</h2>
            <table>
                {testes_html}
            </table>

            <h2>📝 Observações</h2>
            <div class="notes">
                {danos.replace('\\n', '<br>') if danos and danos.strip() else "Nenhuma anomalia visual detetada. Equipamento em condições normais."}
            </div>
        </div>
    </body>
    </html>
    """
    
    safe_serial = "".join([c for c in sys_info.get('serial', 'SN') if c.isalnum()]).strip()
    if not safe_serial: safe_serial = "TV"
    default_filename = f"TV_{safe_serial}.html"
    
    file_path = filedialog.asksaveasfilename(
        defaultextension=".html",
        filetypes=[("Ficheiros HTML", "*.html")],
        initialfile=default_filename,
        title="Guardar Relatório Como..."
    )
    
    if not file_path: return

    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(html_content)
            
        webbrowser.open('file://' + os.path.realpath(file_path))
        
        # Salvar Excel
        if guardar_em_excel_tv(usuario, compra_num, sys_info, portas, testes, danos):
            messagebox.showinfo("Sucesso", "O Teste de TV foi guardado com sucesso!")
        else:
            messagebox.showwarning("Atenção", "Relatório HTML gerado, mas erro ao salvar no Excel.")
            
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao gravar: {e}")

def guardar_em_excel_tv(usuario, compra_num, sys_info, portas, testes, danos):
    try:
        registo = {
            "Data": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
            "Técnico": usuario,
            "Nº Compra": compra_num,
            "Marca/Modelo": (sys_info.get('marca', '') + " " + sys_info.get('modelo', '')).strip() or "Desconhecido",
            "Serial": sys_info.get('serial', 'N/A'),
            "Resolução": sys_info.get('resolution', 'N/A'),
            "Refresh": sys_info.get('refresh_rate', 'N/A'),
            "DisplayPort": portas.get("DisplayPort", 0),
            "HDMI": portas.get("HDMI", 0),
            "DVI": portas.get("DVI", 0),
            "VGA": portas.get("VGA", 0),
            "RS232": portas.get("RS232", 0),
            "USB": portas.get("USB", 0),
            "USB A": portas.get("USB A", 0),
            "USB C": portas.get("USB C", 0),
            "Ecrã / Imagem": "✓" if testes.get("Ecrã / Imagem") else "✗",
            "Touch Screen": "✓" if testes.get("Touch Screen") else "✗",
            "Colunas": "✓" if testes.get("Colunas") else "✗",
            "Cabos / Energia": "✓" if testes.get("Cabos / Energia") else "✗",
            "Botões": "✓" if testes.get("Botões") else "✗",
            "Comando (Remote)": "✓" if testes.get("Comando (Remote)") else "✗",
            "Webcam": "✓" if testes.get("Webcam") else "✗",
            "Notas": danos.strip() if danos.strip() else "Sem observações"
        }
        
        if os.path.exists(EXCEL_FILE_TV):
             df_exist = pd.read_excel(EXCEL_FILE_TV)
             df = pd.concat([df_exist, pd.DataFrame([registo])], ignore_index=True)
        else:
             df = pd.DataFrame([registo])
             
        cols_order = [
            "Data", "Técnico", "Nº Compra", "Marca/Modelo", "Serial", "Resolução", "Refresh",
            "DisplayPort", "HDMI", "DVI", "VGA", "RS232", "USB", "USB A", "USB C",
            "Ecrã / Imagem", "Touch Screen", "Colunas", "Cabos / Energia", "Botões", "Comando (Remote)", "Webcam", "Notas"
        ]
        
        for col in cols_order:
            if col not in df.columns:
                df[col] = "N/A"
                
        df = df[cols_order]
             
        df.to_excel(EXCEL_FILE_TV, index=False, sheet_name="RegistosTV")
        formatar_excel_tv(EXCEL_FILE_TV)
        
        fazer_backup(EXCEL_FILE_TV, "TV")
        return True
    except Exception as e:
        print("Erro Excel TV:", e)
        return False

def formatar_excel_tv(filepath):
    """Aplica formatação moderna ao ficheiro Excel de TVs"""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(left=Side(style='thin', color='E59866'), right=Side(style='thin', color='E59866'),
                             top=Side(style='thin', color='E59866'), bottom=Side(style='thin', color='E59866'))
        
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 25
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 2):
            fill = alt_fill if (row_idx % 2 == 0) else white_fill
            
            for col_idx, cell in enumerate(row, 1):
                cell.fill = fill
                cell.border = thin_border
                
                # Alinhar colunas numéricas de pt/testes
                # 8 ao 15 são portas (DisplayPort até USB C)
                # 16 ao 22 são testes (Ecrã / Imagem até Webcam)
                if col_idx >= 8 and col_idx <= 22: 
                    cell.alignment = center_alignment
                    cell.font = Font(size=12, bold=True)
                    if cell.value == "✓": cell.font = Font(size=12, bold=True, color="00B050")
                    elif cell.value == "✗": cell.font = Font(size=12, bold=True, color="C00000")
                else:
                    cell.alignment = left_alignment
                    cell.font = Font(size=10)
                
                ws.row_dimensions[row_idx].height = 20
        
        column_widths = {
            'A': 16, 'B': 12, 'C': 14, 'D': 30, 'E': 18, 'F': 12, 'G': 10,
            'H': 10, 'I': 8,  'J': 8,  'K': 8,  'L': 8,  'M': 8,  'N': 8,  'O': 8,
            'P': 14, 'Q': 14, 'R': 8,  'S': 14, 'T': 8,  'U': 16, 'V': 10, 'W': 30
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        ws.freeze_panes = "A2"
        wb.save(filepath)
        return True
    except Exception as e:
        print(f"Erro ao formatar Excel TV: {e}")
        return False

if __name__ == "__main__":
    app = App()
    app.mainloop()